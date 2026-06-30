from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Iterable, List, Optional, Tuple, Dict


HEADER_ALIASES = {
    "name": {"功能点", "功能点名称", "功能名称", "模块名称", "建设内容", "名称"},
    "description": {"功能描述", "功能点描述", "描述", "建设内容描述", "主要功能", "说明"},
    "category": {"分类", "功能分类", "类别", "所属模块"},
}


@dataclass(frozen=True)
class ParsedFunctionPoint:
    row_index: int
    name: str
    description: str
    category: Optional[str] = None


def parse_function_points(content: bytes, filename: str) -> List[ParsedFunctionPoint]:
    if filename.lower().endswith(".csv"):
        return _parse_csv(content)
    return _parse_xlsx(content)


def _parse_xlsx(content: bytes) -> List[ParsedFunctionPoint]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    return _parse_rows(rows)


def _parse_csv(content: bytes) -> List[ParsedFunctionPoint]:
    text = content.decode("utf-8-sig")
    rows = list(csv.reader(StringIO(text)))
    return _parse_rows(rows)


def _parse_rows(rows: Iterable[Iterable[object]]) -> List[ParsedFunctionPoint]:
    materialized = [list(row) for row in rows if any(_cell_text(cell) for cell in row)]
    if not materialized:
        return []

    header_index, mapping = _find_header(materialized)
    points: List[ParsedFunctionPoint] = []
    for offset, row in enumerate(materialized[header_index + 1 :], start=header_index + 2):
        name = _value_at(row, mapping.get("name"))
        description = _value_at(row, mapping.get("description"))
        category = _value_at(row, mapping.get("category")) or None
        if not name and description:
            name = description[:40]
        if name:
            points.append(
                ParsedFunctionPoint(
                    row_index=offset,
                    name=name,
                    description=description,
                    category=category,
                )
            )
    return points


def _find_header(rows: List[list[object]]) -> Tuple[int, Dict[str, int]]:
    best_index = 0
    best_mapping: Dict[str, int] = {}
    best_score = -1

    for index, row in enumerate(rows[:10]):
        mapping: Dict[str, int] = {}
        for col_index, cell in enumerate(row):
            normalized = _normalize_header(_cell_text(cell))
            for field, aliases in HEADER_ALIASES.items():
                if normalized in {_normalize_header(alias) for alias in aliases}:
                    mapping[field] = col_index
        score = len(mapping)
        if score > best_score:
            best_index = index
            best_mapping = mapping
            best_score = score

    if "name" not in best_mapping:
        best_mapping["name"] = 0
    if "description" not in best_mapping:
        best_mapping["description"] = 1 if _row_width(rows[best_index]) > 1 else best_mapping["name"]
    return best_index, best_mapping


def _value_at(row: list[object], index: Optional[int]) -> str:
    if index is None or index >= len(row):
        return ""
    return _cell_text(row[index])


def _row_width(row: list[object]) -> int:
    return len([cell for cell in row if _cell_text(cell)])


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: str) -> str:
    return value.replace(" ", "").replace("\n", "").replace("\t", "").strip().lower()
